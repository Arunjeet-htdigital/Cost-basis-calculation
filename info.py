import csv
import re
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
import sys
import pandasql as pds
from collections import defaultdict, deque


HEADER_ANCHOR = "Wallet"


def normalise_columns(columns):
    out = []
    for c in columns:
        c = str(c).strip().lower()
        c = re.sub(r'[\s/()]+', '_', c)
        c = re.sub(r'_+', '_', c).strip('_')
        out.append(c)
    return out


def detect_header_and_clean(df):

    header_row_idx = None
    for i, row in df.iterrows():
        if row.astype(str).str.contains(HEADER_ANCHOR, case=False, na=False).any():
            header_row_idx = i
            break

    if header_row_idx is None:
        return pd.DataFrame()

    header = df.iloc[header_row_idx]
    body = df.iloc[header_row_idx + 1:].reset_index(drop=True)
    body.columns = normalise_columns(header)

    # Drop duplicate column names (keep first) and empty names
    seen = set()
    keep_mask = []
    for c in body.columns:
        if c == '' or c in seen:
            keep_mask.append(False)
        else:
            seen.add(c)
            keep_mask.append(True)
    body = body.loc[:, keep_mask]

    return body


def read_folder(folder_path):
    folder = Path(folder_path)
    files = sorted(folder.glob('*.xlsx'))
    if not files:
        raise FileNotFoundError(f"No XLSX files in {folder}")

    frames = []
    for f in files:
        print(f"Reading {f.name}...")
        wb = load_workbook(f, data_only=True, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()


        raw = pd.DataFrame(rows)
        df = detect_header_and_clean(raw)
        if df.empty:
            print(f"  skip: no header row found in {f.name}")
            continue
        df['source_file'] = f.name
        frames.append(df)
        print(f"  {len(df):,} rows")

    if not frames:
        return pd.DataFrame()

    combined = pd.concat(frames, ignore_index=True)
    print(f"\nCombined: {len(combined):,} rows from {len(frames)} file(s)")
    return combined


def snapshot(folder_path):
    df = read_folder(folder_path)

    df2=df

    query = """
    SELECT b.* from
    (WITH init AS (
        SELECT 
            wallet,asset,
            '2024-01-31 00:00:00' AS timestamp, 
            SUM(quantity) AS quantity, 
            SUM(quantity) AS total_quantity 
        FROM df
        GROUP BY wallet, asset
    ),

    avco AS (
        SELECT 
            wallet, asset,
            SUM(inventory_cost_gbp) * 1.0 / SUM(quantity) AS avco
        FROM df2
        GROUP BY wallet, asset
    )

    SELECT 
        i.*,
        av.avco
    FROM init i
    LEFT JOIN avco av 
        ON i.wallet=av.wallet AND i.asset = av.asset
    ) AS b
    WHERE b.avco is NOT NULL
    """
    res=pds.sqldf(query, locals())
    res["wallet"]=res["wallet"].apply(lambda x:str(x).strip())

    if 'timestamp' in res.columns:
        res['timestamp'] = pd.to_datetime(
        res['timestamp'].astype(str).str.strip(),
        errors='coerce',
        utc=True
    )

    data= defaultdict(lambda: defaultdict(deque))

    for i, row in res.iterrows():
        if row["wallet"]=="Builder-Coinbase-New" and row["asset"]=="ETH":
            data[row["wallet"]][row["asset"]].append((row["timestamp"],row["quantity"], row["total_quantity"],1813))
        else:            
            data[row["wallet"]][row["asset"]].append((row["timestamp"],row["quantity"], row["total_quantity"], row["avco"]))

    return data

folder_snap = sys.argv[1] if len(sys.argv) > 1 else r"C:\Users\Arunjeet\OneDrive - Harris & Trotter LLP\Downloads\test_keeper_snap"
snap = snapshot(folder_snap)
print(snap)


   



