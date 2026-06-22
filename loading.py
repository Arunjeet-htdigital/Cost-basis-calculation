import csv
import re
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook


HEADER_ANCHOR = "hash"


def normalise_columns(columns):
    out = []
    for c in columns:
        c = str(c).strip().lower()
        c = re.sub(r'[\s/()]+', '_', c)
        c = re.sub(r'_+', '_', c).strip('_')
        out.append(c)
    return out


def detect_header_and_clean(df):

    header_row_idx = None
    for i, row in df.iterrows():
        if row.astype(str).str.contains(HEADER_ANCHOR, case=False, na=False).any():
            header_row_idx = i
            break

    if header_row_idx is None:
        return pd.DataFrame()

    header = df.iloc[header_row_idx]
    body = df.iloc[header_row_idx + 1:].reset_index(drop=True)
    body.columns = normalise_columns(header)

    # Drop duplicate column names (keep first) and empty names
    seen = set()
    keep_mask = []
    for c in body.columns:
        if c == '' or c in seen:
            keep_mask.append(False)
        else:
            seen.add(c)
            keep_mask.append(True)
    body = body.loc[:, keep_mask]

    # Parse timestamp from date_utc
    if 'date_utc' in body.columns:
        body['timestamp'] = pd.to_datetime(
        body['date_utc'].astype(str).str.strip(),
        errors='coerce',
        utc=True
    )

    # Keep only these columns (after normalisation)
    KEEP_COLUMNS = {
        'hash_unique_id',
        'source_name',
        'order_type',
        'incoming_asset_unique_symbol',
        'incoming_volume',
        'incoming_asset_price_gbp',
        'outgoing_asset_unique_symbol',
        'outgoing_volume',
        'outgoing_asset_price_gbp',
        'fee_asset_unique_symbol',
        'fee_volume',
        'fee_asset_price_gbp',
        'fee_book_value_gbp',
        'fee_value_gbp',
        'internal_transfer',
        'timestamp',
        'labels',
        'other_parties'}

    #DO THE NECESSARY CHANGE TRANSFORMATIONS.........................
    body = body[[c for c in body.columns if c in KEEP_COLUMNS]]

    print(len(body))

   # body = body[
    #(body['internal_transfer'] == 'NO')]

    print(len(body))   

    #body=body[
    #(~body['labels'].str.lower().str.contains('staking rewards', na=False))]

    print(len(body))

    body.drop_duplicates(subset=['hash_unique_id'], inplace=True)

    # Move timestamp to index 0
    if 'timestamp' in body.columns:
        cols = list(body.columns)
        cols.remove('timestamp')
        cols.insert(0, 'timestamp')
        body = body[cols]

    body.rename(columns={'other_parties': 'address'}, inplace=True)

    return body


def read_folder(folder_path):
    folder = Path(folder_path)
    files = sorted(folder.glob('*.xlsx'))
    if not files:
        raise FileNotFoundError(f"No XLSX files in {folder}")

    frames = []
    for f in files:
        print(f"Reading {f.name}...")
        wb = load_workbook(f, data_only=True, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()


        raw = pd.DataFrame(rows)
        df = detect_header_and_clean(raw)
        if df.empty:
            print(f"  skip: no header row found in {f.name}")
            continue
        df['source_file'] = f.name
        frames.append(df)
        print(f"  {len(df):,} rows")

    if not frames:
        return pd.DataFrame()

    combined = pd.concat(frames, ignore_index=True)
    print(f"\nCombined: {len(combined):,} rows from {len(frames)} file(s)")
    return combined


def run(folder_path):
    df = read_folder(folder_path)

    # Sort ascending by timestamp
    df.sort_values('timestamp', ascending=True, inplace=True)

    # Print everything

    return df

   